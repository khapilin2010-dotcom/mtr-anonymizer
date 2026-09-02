from pathlib import Path
import csv
import fitz
from openpyxl import Workbook, load_workbook

from mtr_core import Anonymizer
from MTR_Obezlichivatel import process_excel, process_pdf


def test_excel_six_fields_and_unicode_path(tmp_path):
    folder = tmp_path / "Русская папка с пробелами"; folder.mkdir()
    src, dst = folder / "вход.xlsx", folder / "результат.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Код Автодокс", "Наименование", "Производитель"])
    ws.append(["", "Клапан Унипол IP66 УХЛ1", ""]); wb.save(src)
    report = process_excel(src, dst, Anonymizer())
    result = load_workbook(dst).active
    headers = [c.value for c in result[1]]
    assert {"Код Автодокс", "Наименование", "Производитель", "Обезличенное наименование", "Статус проверки", "Что именно удалено"} <= set(headers)
    assert "IP66 УХЛ1" in result.cell(2, headers.index("Обезличенное наименование") + 1).value
    removed = result.cell(2, headers.index("Что именно удалено") + 1).value
    assert removed and "Унипол" in removed
    assert report["rows"] == 1 and report["changed"] == 1
    assert report["changed_without_removal_log"] == 0
    assert report["keep_losses"] == 0
    assert report["idempotence_failures"] == 0
    assert report["residual_confirmed"] == 0


def test_text_pdf_physical_redaction(tmp_path):
    src, dst = tmp_path / "in.pdf", tmp_path / "out.pdf"
    doc = fitz.open(); page = doc.new_page(); page.insert_text((72, 72), "Valve UNIPOL IP66 DN50 PN16")
    doc.save(src); doc.close()
    # Latin is not a confirmed Cyrillic rule; use a PDF with a Cyrillic-capable
    # built-in font and verify deletion through the existing service rule.
    doc = fitz.open(); page = doc.new_page(); page.insert_text((72, 72), "Valve sales@example.com IP66 DN50")
    doc.save(src); doc.close()
    process_pdf(src, dst, Anonymizer())
    text = "".join(page.get_text() for page in fitz.open(dst))
    assert "sales@example.com" not in text
    assert "IP66" in text and "DN50" in text


def test_excel_control_audit_has_complete_removal_log(tmp_path):
    with Path("Контроль_200_v15_FINAL.csv").open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle, delimiter=";"))
    src, dst = tmp_path / "control.xlsx", tmp_path / "control_result.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Код Автодокс", "Наименование", "Производитель"])
    for row in rows:
        ws.append([row["Код Autodocs"], row["Исходное"], row["Завод"]])
    wb.save(src)
    report = process_excel(src, dst, Anonymizer())
    assert report["rows"] == 200
    assert report["changed"] == 114
    assert report["green"] == 176 and report["yellow"] == 24
    assert report["changed_without_removal_log"] == 0
    assert report["keep_losses"] == 0
    assert report["idempotence_failures"] == 0
    assert report["residual_confirmed"] == 0
