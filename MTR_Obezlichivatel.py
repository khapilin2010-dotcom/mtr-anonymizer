# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import difflib
import os
import re
import sys
import threading
import traceback
import json
import subprocess
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from mtr_core import Anonymizer, PDF_PROTECTED_RES, keep_signature

APP_TITLE = "Обезличивание МТР"
APP_VERSION = "17.0 RC1"
APP_DIR = Path(os.environ.get("LOCALAPPDATA", Path.home())) / "MTR_Obezlichivatel"
LOG_FILE = APP_DIR / "mtr_v17.log"
CONFIG_FILE = APP_DIR / "config.json"


def resource_path(name: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base / name


def norm_header(value) -> str:
    s = str(value or "").strip().lower().replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    return s


def find_header_row(ws, max_rows: int = 30):
    code_names = {"код autodocs", "код автодокс", "код", "autodocs", "код мтр", "код ресурса", "№ смет/ код ресурса", "№ смет / код ресурса"}
    name_names = {"наименование", "наименование мтр", "мтр", "описание"}
    factory_names = {"завод", "изготовитель", "производитель", "поставщик", "завод/изготовитель/поставщик"}
    anon_names = {"обезличенное наименование", "обезличенное", "результат"}
    status_names = {"статус", "статус обезличивания"}

    for row in range(1, min(ws.max_row, max_rows) + 1):
        vals = [norm_header(ws.cell(row, c).value) for c in range(1, min(ws.max_column, 100) + 1)]
        mapping = {}
        for c, h in enumerate(vals, 1):
            if h in code_names and "code" not in mapping:
                mapping["code"] = c
            if h in name_names and "name" not in mapping:
                mapping["name"] = c
            if h in factory_names and "factory" not in mapping:
                mapping["factory"] = c
            if h in anon_names and "anon" not in mapping:
                mapping["anon"] = c
            if h in status_names and "status" not in mapping:
                mapping["status"] = c
        if "name" in mapping and ("code" in mapping or "factory" in mapping):
            return row, mapping
    raise ValueError(
        "Не найдена строка заголовков. Нужны столбцы «Наименование» и хотя бы один из: «Код Autodocs/Код ресурса» / «Завод»."
    )


def select_excel_sheets(wb):
    """Pick the intended working sheet in estimate/resource workbooks.

    Priority requested for production files: exact "Готово", otherwise exact
    "Выборка оборудования". A helper sheet such as
    "выборка_оборудования_все" is intentionally excluded. For ordinary
    input workbooks, fall back to all sheets and let header detection decide.
    """
    def title_key(title):
        s = str(title or "").strip().lower().replace("ё", "е").replace("_", " ")
        return re.sub(r"\s+", " ", s)

    ready = [ws for ws in wb.worksheets if title_key(ws.title) == "готово"]
    if ready:
        return ready
    equipment = [ws for ws in wb.worksheets if title_key(ws.title) == "выборка оборудования"]
    if equipment:
        return equipment
    return list(wb.worksheets)


def process_excel(src: Path, dst: Path, az: Anonymizer, progress=None):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError as e:
        raise RuntimeError("Для Excel требуется пакет openpyxl. Запустите install_and_run.bat.") from e

    keep_vba = src.suffix.lower() == ".xlsm"
    wb = load_workbook(src, keep_vba=keep_vba)
    report = {"rows": 0, "changed": 0, "green": 0, "yellow": 0, "sheets": 0,
              "keep_losses": 0, "idempotence_failures": 0, "residual_confirmed": 0,
              "changed_without_removal_log": 0}

    green_fill = PatternFill("solid", fgColor="D9EAD3")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    processed_any = False
    for ws in select_excel_sheets(wb):
        try:
            header_row, cols = find_header_row(ws)
        except ValueError:
            continue
        processed_any = True
        report["sheets"] += 1

        # Existing columns and formatting are retained; only missing report
        # fields are appended to the right of the source table.
        factory_col = cols.get("factory")
        if not factory_col:
            factory_col = ws.max_column + 1
            ws.cell(header_row, factory_col).value = "Выявленный завод/производитель"
        anon_col = cols.get("anon")
        if not anon_col:
            anon_col = ws.max_column + 1
            ws.cell(header_row, anon_col).value = "Обезличенное наименование"
        status_col = cols.get("status")
        if not status_col:
            status_col = max(ws.max_column, anon_col) + 1
            ws.cell(header_row, status_col).value = "Статус проверки"
        removed_col = max(ws.max_column, status_col) + 1
        ws.cell(header_row, removed_col).value = "Что именно удалено"

        for c in (factory_col, anon_col, status_col, removed_col):
            cell = ws.cell(header_row, c)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        data_rows = max(ws.max_row - header_row, 0)
        for i, r in enumerate(range(header_row + 1, ws.max_row + 1), 1):
            name_cell = ws.cell(r, cols["name"])
            name = str(name_cell.value or "").strip()
            # Formula/error rows such as #DIV/0! are workbook calculations,
            # not MTR positions and must not enter anonymization statistics.
            if not name or name_cell.data_type == "e" or re.fullmatch(r"#[A-Z0-9/?!._-]+", name, re.I):
                continue
            code = str(ws.cell(r, cols.get("code", 0)).value or "").strip() if cols.get("code") else ""
            factory = str(ws.cell(r, cols.get("factory", 0)).value or "").strip() if cols.get("factory") else ""
            # Estimate workbooks often have a second header row with column
            # numbers (1/2/3/4). It is layout metadata, not an MTR position.
            if r <= header_row + 5 and code in {"1", "2"} and name in {"3", "4"}:
                continue
            result = az.anonymize(name, code, factory)
            # Built-in regression guards: they do not alter the result.
            before_keep = keep_signature(name)
            after_keep = keep_signature(result["text"])
            if before_keep != after_keep:
                report["keep_losses"] += 1
            repeat = az.anonymize(result["text"], code, factory)
            if repeat["text"] != result["text"]:
                report["idempotence_failures"] += 1
            if az.redaction_spans(result["text"], code, factory):
                report["residual_confirmed"] += 1

            ws.cell(r, anon_col).value = result["text"]
            ws.cell(r, status_col).value = result["status"]
            ws.cell(r, factory_col).value = result["factory"] or factory
            removal_log = "; ".join(result["removed"])
            ws.cell(r, removed_col).value = removal_log
            if result["changed"] and not removal_log:
                report["changed_without_removal_log"] += 1
            fill = green_fill if result["status"] == "ЗЕЛЁНЫЙ" else yellow_fill
            ws.cell(r, status_col).fill = fill
            if cols.get("code"):
                ws.cell(r, cols["code"]).fill = fill

            report["rows"] += 1
            report["changed"] += int(result["changed"])
            if result["status"] == "ЗЕЛЁНЫЙ":
                report["green"] += 1
            else:
                report["yellow"] += 1

            if progress and (i % 100 == 0 or i == data_rows):
                progress(f"{ws.title}: обработано {i} из {data_rows}")

        ws.column_dimensions[ws.cell(header_row, anon_col).column_letter].width = min(max(ws.column_dimensions[ws.cell(header_row, cols["name"]).column_letter].width or 25, 40), 100)
        ws.column_dimensions[ws.cell(header_row, status_col).column_letter].width = 15
        ws.column_dimensions[ws.cell(header_row, removed_col).column_letter].width = 45

    if not processed_any:
        raise ValueError("Ни на одном листе не найдены подходящие столбцы для обработки.")

    wb.save(dst)
    return report


def process_csv(src: Path, dst: Path, az: Anonymizer, progress=None):
    raw = src.read_bytes()
    encoding = "utf-8-sig"
    try:
        text = raw.decode(encoding)
    except UnicodeDecodeError:
        encoding = "cp1251"
        text = raw.decode(encoding)
    sample = text[:5000]
    dialect = csv.Sniffer().sniff(sample, delimiters=";,\t,")
    rows = list(csv.reader(text.splitlines(), dialect))
    if not rows:
        raise ValueError("CSV пуст.")

    header_idx = None
    mapping = None
    for idx, row in enumerate(rows[:30]):
        heads = [norm_header(x) for x in row]
        m = {}
        for c, h in enumerate(heads):
            if h in {"код autodocs", "код автодокс", "код", "autodocs", "код мтр"} and "code" not in m:
                m["code"] = c
            if h in {"наименование", "наименование мтр", "мтр", "описание"} and "name" not in m:
                m["name"] = c
            if h in {"завод", "изготовитель", "производитель", "поставщик", "завод/изготовитель/поставщик"} and "factory" not in m:
                m["factory"] = c
        if "name" in m and ("code" in m or "factory" in m):
            header_idx, mapping = idx, m
            break
    if mapping is None:
        raise ValueError("В CSV не найдены столбцы «Наименование» и «Код Autodocs»/«Завод».")

    header = rows[header_idx]
    anon_col = len(header)
    status_col = anon_col + 1
    factory_out_col = mapping.get("factory")
    if factory_out_col is None:
        factory_out_col = len(header)
        header.append("Выявленный завод/производитель")
    anon_col = len(header)
    header.extend(["Обезличенное наименование", "Статус проверки", "Что именно удалено"])
    status_col, removed_col = anon_col + 1, anon_col + 2
    max_len = len(header)
    report = {"rows": 0, "changed": 0, "green": 0, "yellow": 0, "sheets": 1}

    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        row.extend([""] * (max_len - len(row)))
        name = row[mapping["name"]].strip() if mapping["name"] < len(row) else ""
        if not name:
            continue
        code = row[mapping["code"]].strip() if mapping.get("code") is not None and mapping["code"] < len(row) else ""
        factory = row[mapping["factory"]].strip() if mapping.get("factory") is not None and mapping["factory"] < len(row) else ""
        result = az.anonymize(name, code, factory)
        row[anon_col] = result["text"]
        row[status_col] = result["status"]
        row[factory_out_col] = result["factory"] or factory
        row[removed_col] = "; ".join(result["removed"])
        report["rows"] += 1
        report["changed"] += int(result["changed"])
        report["green"] += int(result["status"] == "ЗЕЛЁНЫЙ")
        report["yellow"] += int(result["status"] == "ЖЁЛТЫЙ")
        if progress and i % 200 == 0:
            progress(f"CSV: обработано {i - header_idx} строк")

    with open(dst, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, dialect=dialect)
        writer.writerows(rows)
    return report


def process_xls(src: Path, dst: Path, az: Anonymizer, progress=None):
    """Process legacy BIFF .xls while retaining the source workbook layout."""
    try:
        import xlrd
        from xlutils.copy import copy as copy_xls
        import xlwt
    except ImportError as exc:
        raise RuntimeError("Для старого XLS требуются xlrd, xlwt и xlutils.") from exc
    source = xlrd.open_workbook(str(src), formatting_info=True)
    output = copy_xls(source)
    report = {"rows": 0, "changed": 0, "green": 0, "yellow": 0, "sheets": 0}
    green = xlwt.easyxf("pattern: pattern solid, fore_colour light_green;")
    yellow = xlwt.easyxf("pattern: pattern solid, fore_colour light_yellow;")
    for sheet_index, sheet in enumerate(source.sheets()):
        header_row = mapping = None
        for row in range(min(sheet.nrows, 30)):
            heads = [norm_header(sheet.cell_value(row, col)) for col in range(sheet.ncols)]
            found = {}
            for col, heading in enumerate(heads):
                if heading in {"код autodocs", "код автодокс", "код", "autodocs", "код мтр", "код ресурса"}: found.setdefault("code", col)
                if heading in {"наименование", "наименование мтр", "мтр", "описание"}: found.setdefault("name", col)
                if heading in {"завод", "изготовитель", "производитель", "поставщик", "завод/изготовитель/поставщик"}: found.setdefault("factory", col)
            if "name" in found and ("code" in found or "factory" in found):
                header_row, mapping = row, found
                break
        if mapping is None:
            continue
        report["sheets"] += 1
        writable = output.get_sheet(sheet_index)
        next_col = sheet.ncols
        factory_col = mapping.get("factory")
        if factory_col is None:
            factory_col = next_col; next_col += 1
            writable.write(header_row, factory_col, "Выявленный завод/производитель")
        anon_col, status_col, removed_col = next_col, next_col + 1, next_col + 2
        for col, value in ((anon_col, "Обезличенное наименование"), (status_col, "Статус проверки"), (removed_col, "Что именно удалено")):
            writable.write(header_row, col, value)
        for row in range(header_row + 1, sheet.nrows):
            name = str(sheet.cell_value(row, mapping["name"])).strip()
            if not name:
                continue
            code = str(sheet.cell_value(row, mapping["code"])).strip() if "code" in mapping else ""
            factory = str(sheet.cell_value(row, mapping["factory"])).strip() if "factory" in mapping else ""
            result = az.anonymize(name, code, factory)
            style = green if result["status"] == "ЗЕЛЁНЫЙ" else yellow
            writable.write(row, factory_col, result["factory"] or factory)
            writable.write(row, anon_col, result["text"])
            writable.write(row, status_col, result["status"], style)
            writable.write(row, removed_col, "; ".join(result["removed"]))
            report["rows"] += 1; report["changed"] += int(result["changed"])
            report["green"] += int(result["status"] == "ЗЕЛЁНЫЙ"); report["yellow"] += int(result["status"] == "ЖЁЛТЫЙ")
        if progress: progress(f"XLS: обработан лист {sheet.name}")
    if not report["sheets"]:
        raise ValueError("Ни на одном листе XLS не найдены подходящие столбцы.")
    output.save(str(dst))
    return report


def _block_code(text: str, az: Anonymizer) -> str:
    # Conservative: a registry number is treated as Autodocs code only when it
    # is explicitly labelled or appears at the beginning of a text block.
    text = str(text or "")
    labelled = re.search(r"(?i)код(?:\s+autodocs|\s+автодокс)?\s*[:№-]?\s*(-?\d{3,9})", text)
    if labelled and labelled.group(1) in az.registry:
        return labelled.group(1)
    head = text.lstrip()[:80]
    for token in re.findall(r"(?<!\d)-?\d{3,9}(?!\d)", head):
        if token in az.registry:
            return token
    return ""


def _same_row_code(block_rect, code_blocks):
    best = None
    cy = (block_rect.y0 + block_rect.y1) / 2
    for rect, code in code_blocks:
        overlap = max(0.0, min(block_rect.y1, rect.y1) - max(block_rect.y0, rect.y0))
        min_h = max(1.0, min(block_rect.height, rect.height))
        row_like = overlap / min_h >= 0.45 or abs(cy - (rect.y0 + rect.y1) / 2) <= 7.0
        if not row_like:
            continue
        dist = abs(cy - (rect.y0 + rect.y1) / 2) + 0.001 * abs(block_rect.x0 - rect.x0)
        if best is None or dist < best[0]:
            best = (dist, code)
    return best[1] if best else ""


def _same_row_factory(block_rect, factory_blocks):
    """Return a confirmed manufacturer context from another block on the same visual row."""
    best = None
    cy = (block_rect.y0 + block_rect.y1) / 2
    for rect, factory, inn in factory_blocks:
        overlap = max(0.0, min(block_rect.y1, rect.y1) - max(block_rect.y0, rect.y0))
        min_h = max(1.0, min(block_rect.height, rect.height))
        row_like = overlap / min_h >= 0.45 or abs(cy - (rect.y0 + rect.y1) / 2) <= 7.0
        if not row_like:
            continue
        dist = abs(cy - (rect.y0 + rect.y1) / 2) + 0.001 * abs(block_rect.x0 - rect.x0)
        if best is None or dist < best[0]:
            best = (dist, factory, inn)
    return (best[1], best[2]) if best else ("", "")


def _raw_block_text_map(block):
    """Reconstruct one PyMuPDF raw text block and map every char to its bbox/line."""
    parts = []
    cmap = []
    line_no = 0
    lines = block.get("lines", [])
    for li, line in enumerate(lines):
        for span in line.get("spans", []):
            for ch in span.get("chars", []):
                parts.append(ch.get("c", ""))
                cmap.append((line_no, tuple(ch.get("bbox", (0, 0, 0, 0)))))
        if li != len(lines) - 1:
            parts.append("\n")
            cmap.append(None)
        line_no += 1
    return "".join(parts), cmap


def _tight_redaction_rects(text: str, cmap, start: int, end: int, fitz):
    """Turn a character span into thin per-line redaction strips.

    Many engineering PDFs use font bboxes taller than the physical line pitch.
    Full-height redaction rectangles therefore touch the line above/below and
    PyMuPDF may delete neighbouring technical text.  A thin strip through the
    target glyphs still removes the text object, but cannot intersect adjacent
    lines.
    """
    groups = {}
    for i in range(max(0, start), min(end, len(cmap))):
        item = cmap[i]
        if item is None or not text[i].strip():
            continue
        line_no, bbox = item
        groups.setdefault(line_no, []).append(bbox)

    rects = []
    for line_no in sorted(groups):
        boxes = groups[line_no]
        x0 = min(b[0] for b in boxes)
        y0 = min(b[1] for b in boxes)
        x1 = max(b[2] for b in boxes)
        y1 = max(b[3] for b in boxes)
        h = max(0.5, y1 - y0)
        # Middle-lower band of the target glyph bbox. For the 10 pt Arial used
        # in the supplied specifications this sits fully between neighbouring
        # line bboxes; the ratio also behaves well for other common fonts.
        sy0 = y0 + h * 0.55
        sy1 = y0 + h * 0.62
        if sy1 <= sy0:
            sy1 = sy0 + 0.8
        rects.append(fitz.Rect(x0, sy0, x1, sy1))
    return rects


def _header_key(value) -> str:
    return re.sub(r"[^a-zа-яё0-9]+", " ", str(value or "").casefold()).strip()


def _ocr_word_matches(value: str, expected: str) -> bool:
    """Tolerant header match for predictable OCR glyph substitutions."""
    value = _header_key(value).replace("ё", "е")
    expected = expected.casefold().replace("ё", "е")
    if value == expected:
        return True
    # Common Latin/Cyrillic OCR mixtures are compared in one visual alphabet.
    visual = str.maketrans({"a": "а", "b": "в", "c": "с", "e": "е", "k": "к",
                            "m": "м", "h": "н", "o": "о", "p": "р", "t": "т", "x": "х"})
    value = value.translate(visual)
    return len(value) >= 2 and difflib.SequenceMatcher(None, value, expected).ratio() >= 0.72


def _raster_table_boundaries(page, fitz):
    """Detect long table grid lines in an image-only page.

    OCR spelling is intentionally irrelevant here. A line must occupy over
    half of the page in the other dimension, which excludes glyph strokes and
    leaves the structural grid of specification tables.
    """
    scale = 1.5
    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale),
                          colorspace=fitz.csGRAY, alpha=False)
    samples = memoryview(pix.samples)
    width, height = pix.width, pix.height
    dark = 160
    vertical_pixels = []
    for x in range(width):
        count = sum(samples[y * width + x] < dark for y in range(height))
        if count >= height * 0.50:
            vertical_pixels.append(x)
    horizontal_pixels = []
    for y in range(height):
        row = samples[y * width:(y + 1) * width]
        if sum(value < dark for value in row) >= width * 0.50:
            horizontal_pixels.append(y)

    def clusters(values):
        groups = []
        for value in values:
            if groups and value <= groups[-1][-1] + 2:
                groups[-1].append(value)
            else:
                groups.append([value])
        return [sum(group) / len(group) / scale for group in groups]

    return clusters(vertical_pixels), clusters(horizontal_pixels)


def _table_redaction_zones(page, fitz, az, textpage=None):
    """Return immutable product-code cells and complete supplier text areas.

    Native PDFs use detected table cells. OCR pages fall back to header-word
    geometry and redact complete supplier lines constrained to that column.
    Every returned supplier rectangle is strictly inside its column so table
    borders and the adjacent product-code column cannot be touched.
    """
    code_cells, supplier_areas, code_values = [], [], []
    diagnostics = {"header_words": [], "vertical_boundaries": [],
                   "horizontal_boundaries": [], "code_column_zone": None,
                   "supplier_column_zone": None, "cell_rects": [],
                   "allowed_delete_zones": [], "absolute_keep_zones": [],
                   "reason": "not evaluated"}
    try:
        tables = page.find_tables().tables
    except Exception:
        tables = []
    for table in tables:
        matrix = table.extract()
        for header_row, values in enumerate(matrix[:3]):
            keys = [_header_key(value) for value in values]
            code_col = next((i for i, key in enumerate(keys) if "код продукции" in key), None)
            supplier_col = next((i for i, key in enumerate(keys) if "поставщик" in key), None)
            if code_col is None or supplier_col is None:
                continue
            description_cols = [i for i, key in enumerate(keys) if
                                "наименование" in key or "техническ" in key or
                                "тип марка" in key]
            for row_no in range(header_row + 1, table.row_count):
                for col_no, raw_cell in enumerate(table.rows[row_no].cells):
                    if not raw_cell:
                        continue
                    cell = fitz.Rect(raw_cell)
                    diagnostics["cell_rects"].append({
                        "row": row_no, "column": col_no, "rect": tuple(cell)
                    })
                    if col_no in description_cols:
                        diagnostics["allowed_delete_zones"].append({
                            "role": "description", "row": row_no, "column": col_no,
                            "rect": tuple(fitz.Rect(cell.x0 + 0.5, cell.y0 + 0.5,
                                                    cell.x1 - 0.5, cell.y1 - 0.5)),
                        })
                code_cell = table.rows[row_no].cells[code_col]
                supplier_cell = table.rows[row_no].cells[supplier_col]
                if code_cell:
                    code_rect = fitz.Rect(code_cell)
                    code_cells.append(code_rect)
                    code_values.append(page.get_textbox(code_rect).strip())
                    diagnostics["absolute_keep_zones"].append({
                        "role": "product_code", "row": row_no, "column": code_col,
                        "rect": tuple(code_rect),
                    })
                if supplier_cell:
                    cell_rect = fitz.Rect(supplier_cell)
                    supplier_text = page.get_textbox(cell_rect).strip()
                    _identified, supplier_inn = az.identify_in_text(supplier_text)
                    supplier_org = bool(supplier_inn or re.search(
                        r'(?i)\b(?:ООО|АО|ЗАО|ОАО|ПАО|НПО|НПП|ФГУП|ИНН|завод|производитель|поставщик)\b',
                        supplier_text,
                    ))
                    if supplier_text and supplier_org:
                        # Keep a small inset to preserve vector grid lines.
                        delete_rect = fitz.Rect(
                            cell_rect.x0 + 0.8, cell_rect.y0 + 0.8,
                            cell_rect.x1 - 0.8, cell_rect.y1 - 0.8,
                        )
                        supplier_areas.append(delete_rect)
                        diagnostics["allowed_delete_zones"].append({
                            "role": "supplier", "row": row_no,
                            "column": supplier_col, "rect": tuple(delete_rect),
                        })
            diagnostics["reason"] = "native vector table"
            return code_cells, supplier_areas, code_values, "native-table", diagnostics

    # OCR fallback: derive column bands from the recognized table header.
    if textpage is not None:
        words = page.get_text("words", textpage=textpage, sort=True)
        normalized = [(fitz.Rect(*word[:4]), _header_key(word[4])) for word in words]
        diagnostics["header_words"] = [
            {"text": word[4], "bbox": tuple(word[:4])}
            for word in words if word[1] < page.rect.height * 0.30
        ]
        verticals, horizontals = _raster_table_boundaries(page, fitz)
        diagnostics["vertical_boundaries"] = verticals
        diagnostics["horizontal_boundaries"] = horizontals

        # A standard СО grid has nine columns. Once ten long vertical borders
        # and header/data horizontal borders are visible, their geometry is a
        # stronger signal than potentially damaged OCR header spelling.
        if len(verticals) >= 10 and len(horizontals) >= 3:
            grid_x = verticals[:10]
            grid_y = horizontals
            code_x0, code_x1 = grid_x[3], grid_x[4]
            supplier_x0, supplier_x1 = grid_x[4], grid_x[5]
            header_bottom = grid_y[1]
            diagnostics["code_column_zone"] = (code_x0, header_bottom, code_x1, grid_y[-1])
            diagnostics["supplier_column_zone"] = (supplier_x0, header_bottom,
                                                     supplier_x1, grid_y[-1])
            diagnostics["supplier_source_words"] = [
                {"text": word, "bbox": tuple(rect),
                 "distance_from_code_boundary": rect.x0 - code_x1}
                for rect, word in normalized
                if rect.y0 >= header_bottom and
                rect.intersects(fitz.Rect(supplier_x0, header_bottom,
                                           supplier_x1, grid_y[-1]))
            ]
            for top, bottom in zip(grid_y[1:-1], grid_y[2:]):
                row_no = len(code_cells) + 1
                for col_no, (left, right) in enumerate(zip(grid_x, grid_x[1:])):
                    cell = fitz.Rect(left, top, right, bottom)
                    diagnostics["cell_rects"].append({
                        "row": row_no, "column": col_no, "rect": tuple(cell)
                    })
                    if col_no in (1, 2):
                        diagnostics["allowed_delete_zones"].append({
                            "role": "description", "row": row_no, "column": col_no,
                            "rect": tuple(fitz.Rect(left + 1.0, top + 2.0,
                                                    right - 1.0, bottom - 2.0)),
                        })
                code_rect = fitz.Rect(code_x0, top, code_x1, bottom)
                code_cells.append(code_rect)
                diagnostics["absolute_keep_zones"].append({
                    "role": "product_code", "row": row_no, "column": 3,
                    "rect": tuple(code_rect),
                })
                code_values.append(" ".join(
                    word for rect, word in normalized if rect.intersects(code_rect)
                ))
                # Clear the complete supplier cell while retaining its grid.
                # Start just beyond the raster grid stroke. The asymmetric
                # code KEEP guard below protects the border itself; using a
                # second two-point inset here would leave the first supplier
                # glyph partially outside physical image redaction. Horizontal
                # borders have room for a two-point guard before text begins.
                supplier_rect = fitz.Rect(
                    supplier_x0 + 1.0, top + 2.0,
                    supplier_x1 - 0.8, bottom - 2.0,
                )
                supplier_areas.append(supplier_rect)
                diagnostics["allowed_delete_zones"].append({
                    "role": "supplier", "row": row_no, "column": 4,
                    "rect": tuple(supplier_rect),
                })
            diagnostics["reason"] = "nine-column raster grid"
            return code_cells, supplier_areas, code_values, "ocr-grid-columns", diagnostics

        code_word = next(((rect, word) for rect, word in normalized if _ocr_word_matches(word, "код")), None)
        product_word = next(((rect, word) for rect, word in normalized if _ocr_word_matches(word, "продукции")), None)
        supplier_word = next(((rect, word) for rect, word in normalized if _ocr_word_matches(word, "поставщик")), None)
        unit_word = next(((rect, word) for rect, word in normalized if _ocr_word_matches(word, "ед")), None)
        type_word = next(((rect, word) for rect, word in normalized if
                          _ocr_word_matches(word, "тип") or _ocr_word_matches(word, "марка")), None)
        if code_word and product_word and supplier_word:
            code_header = code_word[0] | product_word[0]
            supplier_header = supplier_word[0]
            code_center = (code_header.x0 + code_header.x1) / 2
            supplier_center = (supplier_header.x0 + supplier_header.x1) / 2
            left_center = ((type_word[0].x0 + type_word[0].x1) / 2) if type_word else code_header.x0 - (supplier_center - code_center)
            right_center = ((unit_word[0].x0 + unit_word[0].x1) / 2) if unit_word else supplier_header.x1 + (supplier_center - code_center)
            code_left = (left_center + code_center) / 2
            boundary = (code_center + supplier_center) / 2
            supplier_right = (supplier_center + right_center) / 2
            header_bottom = max(code_header.y1, supplier_header.y1)
            code_band = fitz.Rect(code_left, header_bottom, boundary, page.rect.y1)
            code_cells.append(code_band)
            code_values.append(" ".join(word for rect, word in normalized if rect.intersects(code_band)))
            supplier_words = [rect for rect, word in normalized if rect.y0 >= header_bottom and rect.intersects(
                fitz.Rect(boundary, header_bottom, supplier_right, page.rect.y1)
            )]
            # Merge words on the same OCR line, clamped inside supplier bounds.
            lines = []
            for rect in sorted(supplier_words, key=lambda item: (round(item.y0 / 4), item.x0)):
                if lines and abs(lines[-1].y0 - rect.y0) <= 4:
                    lines[-1] |= rect
                else:
                    lines.append(fitz.Rect(rect))
            for rect in lines:
                supplier_areas.append(fitz.Rect(
                    max(boundary + 0.8, rect.x0 - 0.8), rect.y0 - 0.8,
                    min(supplier_right - 0.8, rect.x1 + 0.8), rect.y1 + 0.8,
                ))
            diagnostics["code_column_zone"] = tuple(code_band)
            diagnostics["supplier_column_zone"] = (boundary, header_bottom,
                                                     supplier_right, page.rect.y1)
            diagnostics["reason"] = "OCR header words"
            return code_cells, supplier_areas, code_values, "ocr-columns", diagnostics

        # If OCR damaged a header word beyond fuzzy recognition, recover the
        # same two adjacent columns from strongly typed cell contents. Product
        # codes are 6-9 digits; supplier cells contain a legal form / INN. The
        # computed bands are still clamped to text geometry, never page-wide.
        code_tokens = [rect for rect, word in normalized if re.fullmatch(r"\d{6,9}", word)]
        supplier_tokens = [rect for rect, word in normalized if (
            _ocr_word_matches(word, "ооо") or _ocr_word_matches(word, "ао") or
            _ocr_word_matches(word, "инн") or _ocr_word_matches(word, "завод")
        )]
        if code_tokens and supplier_tokens:
            code_x0 = min(rect.x0 for rect in code_tokens) - 2
            supplier_x0 = min(rect.x0 for rect in supplier_tokens) - 2
            # Reject unrelated number / organization layouts: the supplier
            # column must be immediately to the right of product codes.
            if code_x0 < supplier_x0 and supplier_x0 - code_x0 < page.rect.width * 0.30:
                data_top = min(rect.y0 for rect in code_tokens + supplier_tokens) - 2
                supplier_x1 = min(
                    page.rect.x1,
                    unit_word[0].x0 - 1 if unit_word else
                    supplier_x0 + (supplier_x0 - code_x0) * 1.15,
                )
                code_band = fitz.Rect(code_x0, data_top, supplier_x0, page.rect.y1)
                code_cells.append(code_band)
                code_values.append(" ".join(word for rect, word in normalized if rect.intersects(code_band)))
                supplier_band = fitz.Rect(supplier_x0, data_top, supplier_x1, page.rect.y1)
                supplier_words = [rect for rect, _word in normalized if rect.intersects(supplier_band)]
                lines = []
                for rect in sorted(supplier_words, key=lambda item: (round(item.y0 / 4), item.x0)):
                    if lines and abs(lines[-1].y0 - rect.y0) <= 4:
                        lines[-1] |= rect
                    else:
                        lines.append(fitz.Rect(rect))
                supplier_areas.extend(fitz.Rect(
                    max(supplier_x0 + 0.8, rect.x0 - 0.8), rect.y0 - 0.8,
                    min(supplier_x1 - 0.8, rect.x1 + 0.8), rect.y1 + 0.8,
                ) for rect in lines)
                diagnostics["code_column_zone"] = tuple(code_band)
                diagnostics["supplier_column_zone"] = tuple(supplier_band)
                diagnostics["reason"] = "typed OCR cell contents"
                return code_cells, supplier_areas, code_values, "ocr-content-columns", diagnostics
        diagnostics["reason"] = "no usable grid, headers, or typed adjacent columns"
    else:
        diagnostics["reason"] = "no OCR textpage"
    return code_cells, supplier_areas, code_values, "none", diagnostics


def _outside_protected_columns(rect, protected, fitz):
    """Split a candidate so no returned rectangle intersects a KEEP cell."""
    pieces = [fitz.Rect(rect)]
    for keep in protected:
        next_pieces = []
        for piece in pieces:
            overlap = piece & keep
            if overlap.is_empty or overlap.get_area() <= 0:
                next_pieces.append(piece)
                continue
            if piece.x0 < keep.x0:
                next_pieces.append(fitz.Rect(piece.x0, piece.y0, keep.x0, piece.y1))
            if keep.x1 < piece.x1:
                next_pieces.append(fitz.Rect(keep.x1, piece.y0, piece.x1, piece.y1))
        pieces = next_pieces
    return [piece for piece in pieces if piece.width > 0.4 and piece.height > 0.4]


def _outside_protected_rectangles(rects, protected, fitz):
    """Subtract raster-safe grid guards from redaction rectangles."""
    pieces = [fitz.Rect(rect) for rect in rects]
    for guard in protected:
        guard = fitz.Rect(guard)
        next_pieces = []
        for piece in pieces:
            overlap = piece & guard
            if overlap.is_empty or overlap.get_area() <= 0:
                next_pieces.append(piece)
                continue
            if piece.y0 < overlap.y0:
                next_pieces.append(fitz.Rect(piece.x0, piece.y0,
                                             piece.x1, overlap.y0))
            if overlap.y1 < piece.y1:
                next_pieces.append(fitz.Rect(piece.x0, overlap.y1,
                                             piece.x1, piece.y1))
            if piece.x0 < overlap.x0:
                next_pieces.append(fitz.Rect(piece.x0, overlap.y0,
                                             overlap.x0, overlap.y1))
            if overlap.x1 < piece.x1:
                next_pieces.append(fitz.Rect(overlap.x1, overlap.y0,
                                             piece.x1, overlap.y1))
        pieces = next_pieces
    return [piece for piece in pieces if piece.width > 0.4 and piece.height > 0.4]


def _inside_allowed_delete_zones(candidate, source_bbox, zones, fitz,
                                 roles=("description",)):
    """Intersect a candidate with its source cell's explicit delete zone."""
    if not zones:
        return [fitz.Rect(candidate)]
    source = fitz.Rect(source_bbox)
    center = fitz.Point((source.x0 + source.x1) / 2,
                        (source.y0 + source.y1) / 2)
    matching = []
    for zone in zones:
        if zone.get("role") not in roles:
            continue
        rect = fitz.Rect(zone["rect"])
        if center in rect or (source & rect).get_area() > 0:
            clipped = fitz.Rect(candidate) & rect
            if not clipped.is_empty and clipped.width > 0.4 and clipped.height > 0.4:
                matching.append(clipped)
    return matching


def _rectangle_distance(rect, keep):
    """Shortest page-coordinate distance between two rectangles."""
    dx = max(keep.x0 - rect.x1, rect.x0 - keep.x1, 0.0)
    dy = max(keep.y0 - rect.y1, rect.y0 - keep.y1, 0.0)
    return (dx * dx + dy * dy) ** 0.5


def _distance_to_grid_boundary(rect, grid):
    boundary = grid["boundary"]
    if grid["orientation"] == "horizontal":
        return max(rect.y0 - boundary, boundary - rect.y1, 0.0)
    return max(rect.x0 - boundary, boundary - rect.x1, 0.0)


def _ocr_technical_keep_areas(text_blocks, words, fitz, supplier_zone=None,
                              column_boundaries=None):
    """Map protected technical OCR fragments to their physical glyph boxes."""
    protected = []

    supplier_zone = fitz.Rect(supplier_zone) if supplier_zone else None

    def strong_technical(value):
        return bool(re.search(
            r"(?i)(?:IP\s*\d|(?:УХЛ|ХЛ)\s*\d|\bEx\b|\bDN\s*\d|"
            r"\bPN\s*\d|\bГОСТ\s*\d|\b(?:ТУ|СТО|ТТП|TTP)\s*\d|"
            r"\d+(?:[xх×]\d+)+|\d{1,3}Г\d|(?:ОЛ|OL)\d)", value
        ))

    def add(label, value, boxes):
        if not boxes:
            return
        rect = fitz.Rect(boxes[0])
        for box in boxes[1:]:
            rect |= fitz.Rect(box)
        glyph_rect = fitz.Rect(rect)
        # Windows control runs proved that 2.0 pt still lets outward image-pixel
        # rounding alter IP66 while 2.5 pt is the smallest remaining tested
        # halo that stays inside the real gap to the neighbouring brand glyph.
        raster_guard = 2.5
        rect = fitz.Rect(rect.x0 - raster_guard, rect.y0 - raster_guard,
                         rect.x1 + raster_guard, rect.y1 + raster_guard)
        if supplier_zone:
            glyph_center = (glyph_rect.x0 + glyph_rect.x1) / 2
            in_supplier = supplier_zone.x0 <= glyph_center <= supplier_zone.x1
            if in_supplier and not strong_technical(value):
                return
            # A KEEP originating in the technical columns must never grow into
            # the geometrically known supplier column through neighbour union
            # or antialias padding.
            if not in_supplier and glyph_rect.x1 <= supplier_zone.x0:
                rect.x1 = min(rect.x1, supplier_zone.x0)
        key = tuple(round(number, 2) for number in rect)
        if not any(item["key"] == key for item in protected):
            protected.append({"key": key, "label": label, "text": value,
                              "glyph_rect": glyph_rect, "rect": rect,
                              "raster_guard": raster_guard})

    # Primary mapping uses the exact protected spans already employed by
    # Anonymizer.redaction_spans, but converts them back to OCR line geometry.
    for _block_rect, text, cmap in text_blocks:
        for pattern in PDF_PROTECTED_RES:
            for match in pattern.finditer(text):
                by_line = {}
                for index in range(match.start(), min(match.end(), len(cmap))):
                    item = cmap[index]
                    if item is None or not text[index].strip():
                        continue
                    line, bbox = item
                    by_line.setdefault(line, []).append(bbox)
                for boxes in by_line.values():
                    add("protected_regex", match.group(0), boxes)

    # Word-level backup protects technical tokens even when OCR whitespace or
    # a glyph substitution prevents a multi-character regex match.
    technical_word = re.compile(
        r"(?i)^(?:IP\d+[A-Z]?|(?:УХЛ|ХЛ)\d*|Ex\w*|II[ABC]|T[1-6]|"
        r"DN\d+(?:[.,]\d+)?|PN\d+(?:[.,]\d+)?|\d{1,3}Г\d[А-ЯA-Z]?|"
        r"ГОСТ|ТУ|СТО|ТТП|TTP|\d+(?:[xх×]\d+)+|[A-ZА-Я]{2,}[-./]\S*\d\S*)$"
    )
    normalized_words = [(fitz.Rect(*word[:4]), str(word[4])) for word in words]

    def column_index(rect):
        if not column_boundaries:
            return None
        center = (rect.x0 + rect.x1) / 2
        return next((index for index, (left, right) in enumerate(
            zip(column_boundaries, column_boundaries[1:]))
            if left <= center <= right), None)

    def adjacent_word_items(index, limit=6):
        base_rect = normalized_words[index][0]
        base_column = column_index(base_rect)
        selected, previous = [], base_rect
        for candidate_rect, candidate in normalized_words[index + 1:]:
            if len(selected) >= limit:
                break
            if base_column is not None and column_index(candidate_rect) != base_column:
                continue
            same_line = abs((candidate_rect.y0 + candidate_rect.y1) / 2 -
                            (base_rect.y0 + base_rect.y1) / 2) <= max(
                                base_rect.height, candidate_rect.height)
            gap = candidate_rect.x0 - previous.x1
            if not same_line or gap < -1 or gap > max(40, base_rect.height * 6):
                if selected:
                    break
                continue
            selected.append((candidate_rect, candidate))
            previous = candidate_rect
        return selected

    def adjacent_words(index, limit=3):
        return [rect for rect, _word in adjacent_word_items(index, limit)]

    def ocr_key(value):
        """Canonical OCR spelling used only to classify strong KEEP syntax."""
        return re.sub(r"[^0-9A-ZА-Я+.,/\-]", "", str(value).upper()).translate(
            str.maketrans({"Е": "E", "Х": "X", "С": "C", "В": "B",
                           "А": "A", "Т": "T", "О": "O", "Р": "P"})
        )

    def same_cell_sequence(index, limit=7):
        return [(normalized_words[index][0], normalized_words[index][1]),
                *adjacent_word_items(index, limit - 1)]

    # Explosion-protection marks are expressions, not independent words. OCR
    # commonly splits ``Ex d IIC T6`` and may substitute Cyrillic Е/х/С. Build
    # one KEEP from the concrete glyph boxes in the same row and cell. Stop at
    # the first word outside the grammar, so a neighbouring brand is never
    # pulled into the protected rectangle.
    ex_anchor = re.compile(r"^[012]?EX(?:D|DB|IA|IB|IC|E|M|N|P|Q)?$")
    ex_mode = re.compile(r"^(?:D|DB|IA|IB|IC|E|M|N|P|Q)$")
    ex_group = re.compile(r"^II[ABC]$")
    ex_temperature = re.compile(r"^T[1-6]$")
    ex_level = re.compile(r"^G[ABC]$")
    for index, (_rect, word) in enumerate(normalized_words):
        first = ocr_key(word)
        if not ex_anchor.fullmatch(first):
            continue
        expression = same_cell_sequence(index)
        accepted = [expression[0]]
        seen_group = bool(ex_group.fullmatch(first))
        seen_temperature = bool(ex_temperature.fullmatch(first))
        for box, value in expression[1:]:
            key = ocr_key(value)
            if (len(accepted) == 1 and ex_mode.fullmatch(key)) or ex_group.fullmatch(key):
                accepted.append((box, value)); seen_group |= bool(ex_group.fullmatch(key))
            elif ex_temperature.fullmatch(key) and seen_group:
                accepted.append((box, value)); seen_temperature = True
            elif ex_level.fullmatch(key) and seen_temperature:
                accepted.append((box, value))
            else:
                break
        if seen_group and seen_temperature:
            add("explosion_protection_expression",
                " ".join(value for _box, value in accepted),
                [box for box, _value in accepted])

    for index, (rect, word) in enumerate(normalized_words):
        if not technical_word.fullmatch(word):
            continue
        boxes = [rect]
        # Normative prefixes and Ex / DN / PN expressions commonly span the
        # next words; protect their concrete neighbouring glyph boxes too.
        if re.fullmatch(r"(?i)(?:ГОСТ|ТУ|СТО|ТТП|TTP|DN|PN)", word):
            boxes.extend(adjacent_words(index))
        add("protected_ocr_word", word, boxes)

    # Required phrase is absolute KEEP even when split into OCR words. Start
    # from fuzzy 'Комплектация' and continue through the OL/designation word.
    for index, (_rect, word) in enumerate(normalized_words):
        if not _ocr_word_matches(word, "комплектация"):
            continue
        phrase = []
        phrase_column = column_index(normalized_words[index][0])
        for box, value in normalized_words[index:index + 20]:
            if phrase_column is not None and column_index(box) != phrase_column:
                continue
            phrase.append((box, value))
            if re.search(r"(?i)(?:ОЛ|OL)\d", value):
                break
        if phrase and re.search(r"(?i)(?:ОЛ|OL)\d", phrase[-1][1]):
            add("required_phrase_ol", " ".join(value for _box, value in phrase),
                [box for box, _value in phrase])

    return protected


def process_pdf(src: Path, dst: Path, az: Anonymizer, progress=None):
    try:
        import fitz
    except ImportError as e:
        raise RuntimeError("Для PDF требуется пакет PyMuPDF. Запустите install_and_run.bat.") from e

    doc = fitz.open(src)
    if doc.needs_pass:
        doc.close()
        raise ValueError("PDF защищён паролем. Сначала снимите защиту с файла.")

    report = {"pages": len(doc), "redactions": 0, "blocks": 0, "matches": 0,
              "tu_continuations": 0, "context_models": 0, "ocr_pages": 0,
              "ocr_failed_pages": 0, "review": False, "table_pages": 0,
              "supplier_cells_redacted": 0, "code_column_intersections": 0,
              "prevented_code_column_overlaps": 0,
              "prevented_grid_overlaps": 0,
              "code_column_unchanged": True, "code_column_values": [],
              "redaction_rects": [], "code_keep_rects": [],
              "ocr_table_diagnostics": [],
              "grid_keep_rects": [], "ocr_redaction_diagnostics": [],
              "technical_keep_rects": [], "technical_keep_diagnostics": [],
              "prevented_technical_keep_overlaps": 0,
              "allowed_delete_zones": [], "absolute_keep_zones": [],
              "redactions_outside_allowed_zones": 0,
              "code_column_redaction_distances": [],
              "closest_code_redaction": None}

    # OCR is page-level: native text pages stay untouched, image-only pages are
    # recognized in Russian + English and redacted on the original page image.
    tessdata = None
    for candidate in (
        str(resource_path("tessdata")),
        os.environ.get("TESSDATA_PREFIX", ""),
        r"C:\Program Files\Tesseract-OCR\tessdata",
        r"C:\Program Files (x86)\Tesseract-OCR\tessdata",
        "/usr/share/tesseract-ocr/5/tessdata",
        "/usr/share/tesseract-ocr/4.00/tessdata",
    ):
        if candidate and Path(candidate).exists():
            tessdata = candidate
            break
    pending_tu_continuation = False
    ocr_legal_form_re = re.compile(r"(?i)\b(?:ООО|АО|ЗАО|ОАО|ПАО|НПО|НПП|ФГУП|[ОO]{3}|[ОO][АA][ОO]|З[АA][ОO]|П[АA][ОO]|[АA][ОO]|FGUP)\b")
    for page_no, page in enumerate(doc, 1):
        native_text = page.get_text("text")
        ocr_page = len(native_text.strip()) < 12
        active_textpage = None
        if ocr_page:
            try:
                kwargs = {"language": "rus+eng", "dpi": 300, "full": True}
                if tessdata:
                    kwargs["tessdata"] = tessdata
                textpage = page.get_textpage_ocr(**kwargs)
                active_textpage = textpage
                raw = page.get_text("rawdict", textpage=textpage)
                report["ocr_pages"] += 1
                if progress:
                    progress(f"PDF: OCR rus+eng, страница {page_no} из {len(doc)}")
            except Exception as exc:
                report["ocr_failed_pages"] += 1
                if progress:
                    progress(f"PDF: OCR не выполнен на стр. {page_no}: {exc}")
                raw = {"blocks": []}
        else:
            raw = page.get_text("rawdict")
        text_blocks = []
        code_blocks = []
        factory_blocks = []

        for rb in raw.get("blocks", []):
            if rb.get("type") != 0:
                continue
            text, cmap = _raw_block_text_map(rb)
            if not text.strip():
                continue
            rect = fitz.Rect(*rb.get("bbox", (0, 0, 0, 0)))
            text_blocks.append((rect, text, cmap))
            ccode = _block_code(text, az)
            if ccode:
                code_blocks.append((rect, ccode))
            # Supplier / producer blocks frequently sit in a separate PDF
            # column from the model designation. Cache only confirmed contexts
            # (known INN / alias / verified marker) for same-row propagation.
            _ff, _fi = az.identify_in_text(text)
            if not _fi:
                _mf, _mi = az.identify_pdf_marker(text)
                if _mi:
                    _ff, _fi = _mf, _mi
            if _fi:
                factory_blocks.append((rect, _ff or az.manufacturer_name_by_inn.get(_fi, ""), _fi))

        code_keep_rects, supplier_areas, code_before, table_mode, table_diagnostics = _table_redaction_zones(
            page, fitz, az, active_textpage
        )
        technical_keep = []
        allowed_delete_zones = table_diagnostics.get("allowed_delete_zones", [])
        report["allowed_delete_zones"].extend(
            {**zone, "page": page_no} for zone in allowed_delete_zones
        )
        report["absolute_keep_zones"].extend(
            {**zone, "page": page_no}
            for zone in table_diagnostics.get("absolute_keep_zones", [])
        )
        if ocr_page and active_textpage is not None:
            ocr_words = page.get_text("words", textpage=active_textpage, sort=True)
            technical_keep = _ocr_technical_keep_areas(
                text_blocks, ocr_words, fitz,
                supplier_zone=table_diagnostics.get("supplier_column_zone"),
                column_boundaries=table_diagnostics.get("vertical_boundaries"),
            )
            report["technical_keep_rects"].extend(
                tuple(item["rect"]) for item in technical_keep
            )
            report["technical_keep_diagnostics"].extend({
                "page": page_no, "label": item["label"], "text": item["text"],
                "glyph_bbox": tuple(item["glyph_rect"]),
                "rect": tuple(item["rect"]), "raster_guard": item["raster_guard"],
            } for item in technical_keep)
            report["absolute_keep_zones"].extend({
                "page": page_no, "role": "technical",
                "label": item["label"], "text": item["text"],
                "rect": tuple(item["rect"]),
            } for item in technical_keep)
        grid_guards = []
        if ocr_page and table_mode == "ocr-grid-columns":
            verticals = table_diagnostics.get("vertical_boundaries", [])
            horizontals = table_diagnostics.get("horizontal_boundaries", [])
            if verticals and horizontals:
                # Empirical raster controls require one point around vertical
                # strokes and two around horizontal strokes. These are KEEP
                # regions, not visual overlays: every later OCR rule is split
                # before PDF_REDACT_IMAGE_PIXELS is applied.
                for x in verticals:
                    grid_guards.append({"orientation": "vertical", "boundary": x,
                                        "rect": fitz.Rect(x - 1.0, horizontals[0],
                                                          x + 1.0, horizontals[-1])})
                for y in horizontals:
                    grid_guards.append({"orientation": "horizontal", "boundary": y,
                                        "rect": fitz.Rect(verticals[0], y - 2.0,
                                                          verticals[-1], y + 2.0)})
                table_diagnostics["grid_guards"] = [
                    {**item, "rect": tuple(item["rect"])} for item in grid_guards
                ]
                report["grid_keep_rects"].extend(
                    tuple(item["rect"]) for item in grid_guards
                )
        if ocr_page and code_keep_rects:
            # Run #8 proved that the changed pixels were exclusively on the
            # *left* code-column border (a redaction originating in column 3).
            # Keep two points there. On the supplier side one raster-grid
            # stroke plus raster rounding (1.0 pt, verified by the pixel
            # boundary control) is sufficient and still lets
            # image redaction cover a supplier glyph beginning at cell x + 2.
            left_guard, right_guard = 2.0, 1.0
            code_keep_rects = [fitz.Rect(
                max(page.rect.x0, rect.x0 - left_guard), rect.y0,
                min(page.rect.x1, rect.x1 + right_guard), rect.y1,
            ) for rect in code_keep_rects]
            table_diagnostics["code_keep_guard"] = {
                "left": left_guard, "right": right_guard,
                "reason": "left boundary pixel diff; supplier glyph clearance",
            }
        if ocr_page:
            table_diagnostics["page"] = page_no
            table_diagnostics["mode"] = table_mode
            report["ocr_table_diagnostics"].append(table_diagnostics)
        if table_mode != "none":
            report["table_pages"] += 1
            report["code_column_values"].extend(code_before)
            report["code_keep_rects"].extend([tuple(rect) for rect in code_keep_rects])

        page_rects = []
        for supplier_rect in supplier_areas:
            safe_rects = _inside_allowed_delete_zones(
                supplier_rect, supplier_rect, allowed_delete_zones, fitz,
                roles=("supplier",),
            )
            safe_rects = [piece for rect in safe_rects
                          for piece in _outside_protected_columns(rect, code_keep_rects, fitz)]
            supplier_technical_intersections = [
                {"label": item["label"], "reason": item["label"],
                 "text": item["text"], "glyph_bbox": tuple(item["glyph_rect"]),
                 "keep_bbox": tuple(item["rect"]),
                 "intersection_area": (supplier_rect & item["rect"]).get_area()}
                for item in technical_keep
                if (supplier_rect & item["rect"]).get_area() > 0
            ]
            before_technical_clip = [fitz.Rect(rect) for rect in safe_rects]
            safe_rects = _outside_protected_rectangles(
                safe_rects, [item["rect"] for item in technical_keep], fitz
            )
            if len(safe_rects) != len(before_technical_clip) or any(
                not any(all(abs(a - b) <= 0.01 for a, b in zip(before, after))
                        for after in safe_rects)
                for before in before_technical_clip
            ):
                report["prevented_technical_keep_overlaps"] += 1
                report["review"] = True
            before_grid_clip = [fitz.Rect(rect) for rect in safe_rects]
            safe_rects = _outside_protected_rectangles(
                safe_rects, [item["rect"] for item in grid_guards], fitz
            )
            if len(safe_rects) != len(before_grid_clip) or any(
                not any(all(abs(a - b) <= 0.01 for a, b in zip(before, after))
                        for after in safe_rects)
                for before in before_grid_clip
            ):
                report["prevented_grid_overlaps"] += 1
            if not safe_rects:
                report["review"] = True
                continue
            for safe_rect in safe_rects:
                page.add_redact_annot(safe_rect, fill=(1, 1, 1), cross_out=False)
                page_rects.append(safe_rect)
                report["redaction_rects"].append(tuple(safe_rect))
                near_grid = [
                    {"orientation": item["orientation"], "boundary": item["boundary"],
                     "guard_rect": tuple(item["rect"]),
                     "distance_to_guard": _rectangle_distance(safe_rect, item["rect"]),
                     "distance_to_boundary": _distance_to_grid_boundary(safe_rect, item)}
                    for item in grid_guards
                    if _rectangle_distance(safe_rect, item["rect"]) < 5.0
                ]
                near_technical = [
                    {"label": item["label"], "text": item["text"],
                     "glyph_bbox": tuple(item["glyph_rect"]),
                     "keep_bbox": tuple(item["rect"]),
                     "distance_to_glyph": _rectangle_distance(safe_rect, item["glyph_rect"]),
                     "distance_to_keep": _rectangle_distance(safe_rect, item["rect"])}
                    for item in technical_keep
                    if _rectangle_distance(safe_rect, item["glyph_rect"]) < 5.0
                ]
                report["ocr_redaction_diagnostics"].append({
                    "page": page_no, "label": "supplier_cell",
                    "original_ocr_glyph_bbox": None,
                    "expanded_bbox": tuple(supplier_rect),
                    "final_safe_bbox": tuple(safe_rect),
                    "text_span": "supplier cell",
                    "technical_keep_intersections": supplier_technical_intersections,
                    "near_technical_keep": near_technical,
                    "near_grid": near_grid,
                })
                report["redactions"] += 1
            report["supplier_cells_redacted"] += 1
        # Some specifications split a TU number exactly at a page boundary:
        # page N ends with "ТУ 2531-" and page N+1 starts with
        # "002-53597015-12". The continuation is still part of the identifier
        # and must be removed, but only when the preceding page proves the
        # context.
        pending_for_this_page = pending_tu_continuation

        def _has_unfinished_tu(rect, txt):
            # A TU may be split by the physical page boundary even if table
            # extraction appends unit/quantity cells after the visible text.
            # Only consider lower-page blocks and a TU prefix whose immediate
            # continuation is not another digit in the same block.
            if rect.y0 < page.rect.height * 0.60:
                return False
            for mm in re.finditer(r"(?i)\bТУ\s+[0-9][0-9.\-/]*-", txt):
                tail = txt[mm.end():].lstrip(" \t\r\n")
                if not tail or not tail[:1].isdigit():
                    return True
            return False

        pending_tu_continuation = any(
            _has_unfinished_tu(rect, txt) for rect, txt, _cm in text_blocks
        )

        for block_index, (block_rect, text, cmap) in enumerate(text_blocks):
            report["blocks"] += 1
            code = _block_code(text, az) or _same_row_code(block_rect, code_blocks)
            factory, inn = az.identify_in_text(text)
            if code:
                resolved_factory, resolved_inn = az.resolve_factory(code, "")
                if resolved_inn:
                    factory, inn = resolved_factory, resolved_inn
            if not inn:
                row_factory, row_inn = _same_row_factory(block_rect, factory_blocks)
                if row_inn:
                    factory, inn = row_factory, row_inn
            if not inn:
                marker_factory, marker_inn = az.identify_pdf_marker(text)
                if marker_inn:
                    factory, inn = marker_factory, marker_inn

            spans = az.redaction_spans(text, code=code, factory=factory if inn else "")

            # OCR may interleave neighbouring table columns between the label
            # "ИНН" and its digits (for example: "ИНН м 2 0.074 ... 5904184047").
            # Once the manufacturer is already confirmed, its exact known INN
            # is an unambiguous identifier and must be physically redacted
            # wherever it appears in the same block.
            if inn and re.search(r"(?i)ИНН", text):
                for _inn_m in re.finditer(r"(?<!\d)" + re.escape(str(inn)) + r"(?!\d)", text):
                    spans.append((_inn_m.start(), _inn_m.end(), "known_inn"))

            # Supplier INN / KPP values are often split into a separate raw PDF
            # block on the following visual line. If the immediately preceding
            # block ends with the identifier label and this block contains only
            # the corresponding digits, redact the continuation as service data.
            if block_index > 0:
                _prev_rect, _prev_text, _prev_cmap = text_blocks[block_index - 1]

                # OCR can split the INN digits into the next raw block together
                # with unrelated table cells. Resolve the previous block's
                # manufacturer and redact only that exact known INN in the
                # current block; never redact the surrounding technical text.
                if re.search(r"(?i)ИНН", _prev_text):
                    _prev_code = _block_code(_prev_text, az) or _same_row_code(_prev_rect, code_blocks)
                    _pf, _pi = az.identify_in_text(_prev_text)
                    if _prev_code:
                        _rf, _ri = az.resolve_factory(_prev_code, "")
                        if _ri:
                            _pf, _pi = _rf, _ri
                    if not _pi:
                        _mf, _mi = az.identify_pdf_marker(_prev_text)
                        if _mi:
                            _pf, _pi = _mf, _mi
                    if _pi:
                        for _cross_m in re.finditer(r"(?<!\d)" + re.escape(str(_pi)) + r"(?!\d)", text):
                            spans.append((_cross_m.start(), _cross_m.end(), "inn_cross_block"))

                if re.search(r"(?i)\bИНН\s*$", _prev_text) and re.fullmatch(r"\s*\d{10,12}\s*", text):
                    spans.append((0, len(text), "inn_continuation"))
                elif re.search(r"(?i)\bКПП\s*$", _prev_text) and re.fullmatch(r"\s*\d{9}\s*", text):
                    spans.append((0, len(text), "kpp_continuation"))

                # A supplier name may wrap into the next raw block: e.g.
                # `ООО Калужский` / `электротехнический завод "КВТ", ИНН ...`.
                # If the previous block starts an organization and this block
                # completes it with an INN, both fragments are service data.
                _prev_org = ocr_legal_form_re.search(_prev_text)
                _this_inn = re.search(r"(?i)\bИНН\s*\d{10,12}\b", text)
                _near = abs(block_rect.y0 - _prev_rect.y1) <= 35 and not (block_rect.x1 < _prev_rect.x0 or block_rect.x0 > _prev_rect.x1 + 120)
                if _prev_org and _this_inn and _near:
                    # current continuation through the INN
                    spans.append((0, _this_inn.end(), "supplier_org_continuation"))

            # If this block starts a legal organization and the next block
            # immediately completes it with an INN, redact the organization
            # tail in this block from the legal form onwards.
            _org_here = ocr_legal_form_re.search(text)
            if _org_here and not re.search(r"(?i)\bИНН\b", text) and block_index + 1 < len(text_blocks):
                _next_rect, _next_text, _next_cmap = text_blocks[block_index + 1]
                _near_next = abs(_next_rect.y0 - block_rect.y1) <= 35 and not (_next_rect.x1 < block_rect.x0 or _next_rect.x0 > block_rect.x1 + 120)
                if _near_next and re.search(r"(?i)\bИНН\s*\d{10,12}\b", _next_text):
                    spans.append((_org_here.start(), len(text), "supplier_org_wrapped"))

            # Cross-page TU continuation, proven by the previous page only.
            # Limit this to the upper portion of the page and to a block that
            # consists only of a hyphenated numeric continuation.
            if pending_for_this_page and block_rect.y0 < 170:
                m_cont = re.fullmatch(r"\s*[0-9]{2,4}(?:-[0-9]{2,12}){1,4}\s*", text)
                if m_cont:
                    a, b = m_cont.span()
                    spans.append((a, b, "tu_continuation"))
                    report["tu_continuations"] += 1
                    pending_for_this_page = False

            # Layout-aware PDF fallback for KШГ catalogue designations. In
            # table PDFs the word KШГ and its catalogue number may be separated
            # in extraction order by unit/quantity columns even though they are
            # visually part of one item. If both occur in the same row block,
            # remove the confirmed firm designation and catalogue token while
            # leaving DN/PN/material/temperature text untouched.
            if re.search(r"(?i)(?<![\w])КШГ(?![\w])", text):
                model_pat = re.compile(r"(?<![\w])(?:71|79)\.[0-9]{3}\.[0-9]{3}\.[A-Za-zА-Яа-я]\.[0-9]+(?:\.[0-9]+)?(?![\w])")
                model_hits = list(model_pat.finditer(text))
                if model_hits:
                    for km in re.finditer(r"(?i)(?<![\w])КШГ(?![\w])", text):
                        spans.append((km.start(), km.end(), "context:КШГ"))
                    for mm in model_hits:
                        spans.append((mm.start(), mm.end(), "context:КШГ-model"))
                    report["context_models"] += len(model_hits)

            # De-duplicate and merge after adding contextual spans.
            spans = sorted(spans, key=lambda x: (x[0], x[1]))
            merged_spans = []
            for a, b, label in spans:
                if merged_spans and a <= merged_spans[-1][1]:
                    merged_spans[-1] = (merged_spans[-1][0], max(merged_spans[-1][1], b),
                                        merged_spans[-1][2] + ";" + label)
                else:
                    merged_spans.append((a, b, label))
            spans = merged_spans

            for a, b, _label in spans:
                rects = _tight_redaction_rects(text, cmap, a, b, fitz)
                rect_details = [(rect, fitz.Rect(rect)) for rect in rects]
                if ocr_page and rects:
                    # OCR comes from an image. Thin strips delete a text object,
                    # but not the visible glyph pixels. Expand each matched strip
                    # to the full OCR glyph band before pixel redaction.
                    full_groups = {}
                    for ci in range(max(0, a), min(b, len(cmap))):
                        item = cmap[ci]
                        if item is None or not text[ci].strip():
                            continue
                        ln, bb = item
                        full_groups.setdefault(ln, []).append(bb)
                    rect_details = []
                    for ln in sorted(full_groups):
                        boxes = full_groups[ln]
                        x0=min(x[0] for x in boxes); y0=min(x[1] for x in boxes)
                        x1=max(x[2] for x in boxes); y1=max(x[3] for x in boxes)
                        original_glyph_bbox = fitz.Rect(x0, y0, x1, y1)
                        rect_details.append((
                            fitz.Rect(x0-0.8, y0-0.8, x1+0.8, y1+0.8),
                            original_glyph_bbox,
                        ))
                if not rect_details:
                    continue
                report["matches"] += 1
                for rect, original_glyph_bbox in rect_details:
                    expanded_rect = fitz.Rect(rect)
                    safe_rects = _inside_allowed_delete_zones(
                        rect, original_glyph_bbox, allowed_delete_zones, fitz,
                        roles=("description",),
                    )
                    if table_mode != "none" and not safe_rects:
                        report["review"] = True
                    safe_rects = [piece for candidate in safe_rects
                                  for piece in _outside_protected_columns(
                                      candidate, code_keep_rects, fitz)]
                    intersected_technical = [
                        {"label": item["label"], "text": item["text"],
                         "glyph_bbox": tuple(item["glyph_rect"]),
                         "keep_bbox": tuple(item["rect"]),
                         "intersection_area": (expanded_rect & item["rect"]).get_area()}
                        for item in technical_keep
                        if (expanded_rect & item["rect"]).get_area() > 0
                    ]
                    before_technical_clip = [fitz.Rect(item) for item in safe_rects]
                    safe_rects = _outside_protected_rectangles(
                        safe_rects, [item["rect"] for item in technical_keep], fitz
                    )
                    if len(safe_rects) != len(before_technical_clip) or any(
                        not any(all(abs(a - b) <= 0.01
                                    for a, b in zip(before, after))
                                for after in safe_rects)
                        for before in before_technical_clip
                    ):
                        report["prevented_technical_keep_overlaps"] += 1
                    before_grid_clip = [fitz.Rect(item) for item in safe_rects]
                    safe_rects = _outside_protected_rectangles(
                        safe_rects, [item["rect"] for item in grid_guards], fitz
                    )
                    if len(safe_rects) != len(before_grid_clip) or any(
                        not any(all(abs(a - b) <= 0.01
                                    for a, b in zip(before, after))
                                for after in safe_rects)
                        for before in before_grid_clip
                    ):
                        report["prevented_grid_overlaps"] += 1
                    if len(safe_rects) != 1 or (
                        safe_rects and any(abs(a - b) > 0.01 for a, b in zip(safe_rects[0], rect))
                    ):
                        report["prevented_code_column_overlaps"] += 1
                    if not safe_rects:
                        report["review"] = True
                    for rect in safe_rects:
                    # Avoid duplicate annotations from overlapping rules.
                        if any(
                            abs(rect.x0-r.x0) < 0.5 and abs(rect.y0-r.y0) < 0.5 and
                            abs(rect.x1-r.x1) < 0.5 and abs(rect.y1-r.y1) < 0.5
                            for r in page_rects
                        ):
                            continue
                        page.add_redact_annot(rect, fill=(1, 1, 1), cross_out=False)
                        page_rects.append(rect)
                        report["redaction_rects"].append(tuple(rect))
                        near_grid = [
                            {"orientation": item["orientation"],
                             "boundary": item["boundary"],
                             "guard_rect": tuple(item["rect"]),
                             "distance_to_guard": _rectangle_distance(rect, item["rect"]),
                             "distance_to_boundary": _distance_to_grid_boundary(rect, item)}
                            for item in grid_guards
                            if _rectangle_distance(rect, item["rect"]) < 5.0
                        ]
                        near_technical = [
                            {"label": item["label"], "text": item["text"],
                             "glyph_bbox": tuple(item["glyph_rect"]),
                             "keep_bbox": tuple(item["rect"]),
                             "distance_to_glyph": _rectangle_distance(rect, item["glyph_rect"]),
                             "distance_to_keep": _rectangle_distance(rect, item["rect"])}
                            for item in technical_keep
                            if _rectangle_distance(rect, item["glyph_rect"]) < 5.0
                        ]
                        report["ocr_redaction_diagnostics"].append({
                            "page": page_no, "label": _label,
                            "source_rule": _label,
                            "original_ocr_glyph_bbox": tuple(original_glyph_bbox),
                            "expanded_bbox": tuple(expanded_rect),
                            "final_safe_bbox": tuple(rect),
                            "text_span": text[a:b],
                            "technical_keep_intersections": intersected_technical,
                            "near_technical_keep": near_technical,
                            "near_grid": near_grid,
                        })
                        if any((rect & keep).get_area() > 0 for keep in code_keep_rects):
                            report["code_column_intersections"] += 1
                        report["redactions"] += 1

        if page_rects:
            if table_mode != "none":
                for final_rect in page_rects:
                    contained = any(
                        abs((final_rect & fitz.Rect(zone["rect"])).get_area() -
                            final_rect.get_area()) < 0.01
                        for zone in allowed_delete_zones
                    )
                    if not contained:
                        report["redactions_outside_allowed_zones"] += 1
                        report["review"] = True
            for redaction_rect in page_rects:
                if not code_keep_rects:
                    continue
                nearest_keep = min(code_keep_rects,
                                   key=lambda keep: _rectangle_distance(redaction_rect, keep))
                distance = _rectangle_distance(redaction_rect, nearest_keep)
                item = {"page": page_no, "redaction_rect": tuple(redaction_rect),
                        "nearest_code_keep_rect": tuple(nearest_keep),
                        "distance": distance}
                report["code_column_redaction_distances"].append(item)
                closest = report["closest_code_redaction"]
                if closest is None or distance < closest["distance"]:
                    report["closest_code_redaction"] = item
            try:
                page.apply_redactions(
                    images=(fitz.PDF_REDACT_IMAGE_PIXELS if ocr_page else fitz.PDF_REDACT_IMAGE_NONE),
                    graphics=fitz.PDF_REDACT_LINE_ART_NONE,
                    text=fitz.PDF_REDACT_TEXT_REMOVE,
                )
            except TypeError:
                page.apply_redactions(images=0)
        if code_keep_rects and not ocr_page:
            code_after = [page.get_textbox(rect).strip() for rect in code_keep_rects]
            if code_after != code_before:
                report["code_column_unchanged"] = False
        if progress:
            progress(f"PDF: страница {page_no} из {len(doc)}")

    if report["redactions"] == 0:
        # For a scan, inability to make a safe automatic deletion is REVIEW,
        # not a fatal queue error. Save a separate copy and continue the batch.
        if report["ocr_pages"] or report["ocr_failed_pages"]:
            report["review"] = True
        else:
            report["unchanged"] = True

    doc.save(dst, garbage=4, deflate=True, clean=True)
    doc.close()
    return report


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} {APP_VERSION}")
        self.geometry("1120x700")
        self.minsize(960, 620)
        self.configure(bg="#F4F8FC")
        self.file_paths: list[Path] = []
        self.last_outputs: list[Path] = []
        APP_DIR.mkdir(parents=True, exist_ok=True)
        self.az = Anonymizer(resource_path("mtr_data.json.gz"))
        self._build_style()
        self._build_ui()

    def _build_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("TFrame", background="#F4F8FC")
        style.configure("Card.TFrame", background="#FFFFFF")
        style.configure("TLabel", background="#F4F8FC", foreground="#1F2D3D", font=("Segoe UI", 10))
        style.configure("Title.TLabel", background="#F4F8FC", foreground="#0A5AA8", font=("Segoe UI Semibold", 22))
        style.configure("Sub.TLabel", background="#F4F8FC", foreground="#5C6F82", font=("Segoe UI", 10))
        style.configure("Card.TLabel", background="#FFFFFF", foreground="#1F2D3D", font=("Segoe UI", 10))
        style.configure("Badge.TLabel", background="#E7F2FD", foreground="#0A5AA8", font=("Segoe UI Semibold", 9), padding=(9,4))
        style.configure("Primary.TButton", font=("Segoe UI Semibold", 11), padding=(16, 10), background="#0A6ACB", foreground="white")
        style.map("Primary.TButton", background=[("active", "#075AAE"), ("disabled", "#A9BDD1")])
        style.configure("Secondary.TButton", font=("Segoe UI", 10), padding=(12, 8))
        style.configure("Horizontal.TProgressbar", troughcolor="#DDE8F2", background="#0A6ACB", bordercolor="#DDE8F2", lightcolor="#0A6ACB", darkcolor="#0A6ACB")

    def _build_ui(self):
        outer = ttk.Frame(self, padding=24)
        outer.pack(fill="both", expand=True)

        top = ttk.Frame(outer)
        top.pack(fill="x")
        ttk.Label(top, text=APP_TITLE, style="Title.TLabel").pack(side="left")
        ttk.Label(top, text=f"База {self.az.version} • 138 630 МТР", style="Badge.TLabel").pack(side="right", pady=4)
        ttk.Label(outer, text="Удаление производителя, ТУ, фирменных моделей и артикулов с защитой технических характеристик.", style="Sub.TLabel").pack(anchor="w", pady=(4, 18))

        card = ttk.Frame(outer, style="Card.TFrame", padding=18)
        card.pack(fill="x")
        ttk.Label(card, text="Файл для обработки", style="Card.TLabel", font=("Segoe UI Semibold", 11)).pack(anchor="w")
        self.file_label = ttk.Label(card, text="Файл не выбран", style="Card.TLabel", foreground="#6D7F90")
        self.file_label.pack(anchor="w", pady=(6, 12))
        ttk.Button(card, text="Выбрать путь к Tesseract", command=self.choose_tesseract, style="Secondary.TButton").pack(anchor="w")

        info = ttk.Frame(outer, padding=(0, 16, 0, 8))
        info.pack(fill="x")
        ttk.Label(info, text="Поддерживается: Excel .xlsx/.xlsm, CSV, текстовые и сканированные PDF (OCR RU+EN). Код Autodocs и ОЛ сохраняются. Неуверенные случаи получают жёлтый/REVIEW статус.", style="Sub.TLabel", wraplength=750).pack(anchor="w")

        self.progress = ttk.Progressbar(outer, mode="determinate", maximum=100)
        self.progress.pack(fill="x", pady=(4, 10))

        # Футер резервируется у нижней границы окна, чтобы подпись разработчика
        # оставалась видимой независимо от высоты журнала обработки.
        footer = ttk.Frame(outer)
        footer.pack(side="bottom", fill="x", pady=(12, 0))
        ttk.Label(footer, text="разработал Хапилин Виктор", style="Sub.TLabel").pack(side="left")
        ttk.Label(footer, text=f"v{APP_VERSION}", style="Sub.TLabel").pack(side="right")

        # Required actions stay together at the bottom of the interface.
        btns = ttk.Frame(outer)
        btns.pack(side="bottom", fill="x", pady=(10, 0))
        self.choose_one_btn = ttk.Button(btns, text="Выбрать файл", command=self.choose_one, style="Secondary.TButton")
        self.choose_one_btn.pack(side="left")
        self.choose_btn = ttk.Button(btns, text="Выбрать несколько файлов", command=self.choose_files, style="Secondary.TButton")
        self.choose_btn.pack(side="left", padx=4)
        self.folder_btn = ttk.Button(btns, text="Выбрать папку", command=self.choose_folder, style="Secondary.TButton")
        self.folder_btn.pack(side="left")
        self.run_btn = ttk.Button(btns, text="Начать обработку", command=self.start_processing, style="Primary.TButton", state="disabled")
        self.run_btn.pack(side="left", padx=4)
        self.open_result_btn = ttk.Button(btns, text="Открыть результат", command=self.open_result, style="Secondary.TButton", state="disabled")
        self.open_result_btn.pack(side="left")
        self.open_btn = ttk.Button(btns, text="Открыть папку результатов", command=self.open_output_folder, style="Secondary.TButton", state="disabled")
        self.open_btn.pack(side="left", padx=4)
        ttk.Button(btns, text="Посмотреть журнал", command=self.open_log, style="Secondary.TButton").pack(side="left")

        log_card = ttk.Frame(outer, style="Card.TFrame", padding=12)
        log_card.pack(fill="both", expand=True)
        ttk.Label(log_card, text="Ход обработки", style="Card.TLabel", font=("Segoe UI Semibold", 10)).pack(anchor="w", pady=(0, 6))
        self.log = tk.Text(log_card, height=10, wrap="word", bg="#FFFFFF", fg="#263746", relief="flat", font=("Consolas", 9), padx=4, pady=4)
        self.log.pack(fill="both", expand=True)
        self.log.configure(state="disabled")

    def log_msg(self, msg: str):
        timestamped = f"{datetime.now():%Y-%m-%d %H:%M:%S} {msg.rstrip()}\n"
        with LOG_FILE.open("a", encoding="utf-8") as handle:
            handle.write(timestamped)
        def _write():
            self.log.configure(state="normal")
            self.log.insert("end", msg.rstrip() + "\n")
            self.log.see("end")
            self.log.configure(state="disabled")
        self.after(0, _write)

    def choose_one(self):
        path = filedialog.askopenfilename(title="Выберите файл МТР", filetypes=[("Поддерживаемые файлы", "*.xlsx *.xls *.xlsm *.csv *.pdf")])
        if path: self._set_files([path])

    def choose_tesseract(self):
        folder = filedialog.askdirectory(title="Выберите папку Tesseract-OCR или tessdata")
        if not folder: return
        chosen = Path(folder)
        tessdata = chosen if chosen.name.lower() == "tessdata" else chosen / "tessdata"
        if not (tessdata / "rus.traineddata").exists():
            messagebox.showerror(APP_TITLE, "В выбранной папке не найден tessdata\\rus.traineddata")
            return
        CONFIG_FILE.write_text(json.dumps({"tessdata": str(tessdata)}, ensure_ascii=False), encoding="utf-8")
        os.environ["TESSDATA_PREFIX"] = str(tessdata)
        self.log_msg(f"Выбран Tesseract: {tessdata}")

    def _set_files(self, paths):
        supported = {".xlsx", ".xls", ".xlsm", ".csv", ".pdf"}
        clean = []
        seen = set()
        for p in paths:
            pp = Path(p)
            if pp.suffix.lower() not in supported or pp.stem.lower().endswith("_обезличено"):
                continue
            key = str(pp.resolve()) if pp.exists() else str(pp)
            if key not in seen:
                seen.add(key)
                clean.append(pp)
        self.file_paths = clean
        if not clean:
            self.file_label.configure(text="Файлы не выбраны")
            self.run_btn.configure(state="disabled")
            return
        label = str(clean[0]) if len(clean) == 1 else f"Выбрано файлов: {len(clean)} • {clean[0].parent}"
        self.file_label.configure(text=label)
        self.run_btn.configure(state="normal")
        self.log_msg(f"Выбрано файлов: {len(clean)}")

    def choose_files(self):
        paths = filedialog.askopenfilenames(
            title="Выберите файлы МТР",
            filetypes=[
                ("Поддерживаемые файлы", "*.xlsx *.xls *.xlsm *.csv *.pdf"),
                ("Excel", "*.xlsx *.xls *.xlsm"), ("CSV", "*.csv"), ("PDF", "*.pdf"),
            ],
        )
        if paths:
            self._set_files(paths)

    def choose_folder(self):
        folder = filedialog.askdirectory(title="Выберите папку с файлами МТР")
        if not folder:
            return
        root = Path(folder)
        paths = sorted(p for p in root.iterdir() if p.is_file())
        self._set_files(paths)

    def _set_busy(self, busy: bool):
        def _do():
            self.choose_btn.configure(state="disabled" if busy else "normal")
            self.choose_one_btn.configure(state="disabled" if busy else "normal")
            self.folder_btn.configure(state="disabled" if busy else "normal")
            self.run_btn.configure(state="disabled" if busy or not self.file_paths else "normal")
            if not busy: self.progress.configure(value=0)
        self.after(0, _do)

    def start_processing(self):
        if not self.file_paths:
            return
        self._set_busy(True)
        threading.Thread(target=self._process_worker, daemon=True).start()

    def _process_one(self, src: Path):
        suffix = src.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            dst = src.with_name(src.stem + "_обезличено" + suffix)
            report = process_excel(src, dst, self.az, self.log_msg)
            summary = f"{src.name}: {report['rows']} строк; зелёных {report['green']}; жёлтых {report['yellow']}; изменено {report['changed']}."
        elif suffix == ".xls":
            dst = src.with_name(src.stem + "_обезличено.xls")
            report = process_xls(src, dst, self.az, self.log_msg)
            summary = f"{src.name}: {report['rows']} строк; зелёных {report['green']}; жёлтых {report['yellow']}; изменено {report['changed']}."
        elif suffix == ".csv":
            dst = src.with_name(src.stem + "_обезличено.csv")
            report = process_csv(src, dst, self.az, self.log_msg)
            summary = f"{src.name}: {report['rows']} строк; зелёных {report['green']}; жёлтых {report['yellow']}; изменено {report['changed']}."
        elif suffix == ".pdf":
            dst = src.with_name(src.stem + "_обезличено.pdf")
            report = process_pdf(src, dst, self.az, self.log_msg)
            tag = "REVIEW" if report.get("review") else "ГОТОВО"
            summary = (f"{src.name}: {tag}; {report['pages']} стр.; удалений {report['redactions']}; "
                       f"OCR-страниц {report.get('ocr_pages', 0)}; OCR-ошибок {report.get('ocr_failed_pages', 0)}.")
        else:
            raise ValueError("Формат не поддерживается.")
        return dst, report, summary

    def _process_worker(self):
        ok = review = errors = 0
        outputs = []
        details = []
        total = len(self.file_paths)
        for idx, src in enumerate(list(self.file_paths), 1):
            try:
                self.log_msg(f"[{idx}/{total}] {src.name}")
                dst, report, summary = self._process_one(src)
                outputs.append(dst)
                if report.get("review"):
                    review += 1
                else:
                    ok += 1
                details.append(summary)
                self.log_msg(summary)
            except Exception as e:
                errors += 1
                details.append(f"{src.name}: ОШИБКА — {e}")
                self.log_msg(f"{src.name}: ОШИБКА — {e}")
                # Critical v15 behavior: one file never stops the remaining queue.
                continue
            finally:
                self.after(0, lambda value=idx * 100 / total: self.progress.configure(value=value))

        self.last_outputs = outputs
        if outputs:
            self.after(0, lambda: (self.open_btn.configure(state="normal"), self.open_result_btn.configure(state="normal")))
        summary = f"Пакет завершён: успешно {ok}; REVIEW {review}; ошибок {errors}; всего {total}."
        self.log_msg(summary)
        self.after(0, lambda: messagebox.showinfo(APP_TITLE, summary))
        self._set_busy(False)

    @staticmethod
    def _open_path(path: Path):
        if sys.platform.startswith("win"): os.startfile(str(path))  # type: ignore[attr-defined]
        elif sys.platform == "darwin": subprocess.Popen(["open", str(path)])
        else: subprocess.Popen(["xdg-open", str(path)])

    def open_result(self):
        if self.last_outputs:
            try: self._open_path(self.last_outputs[-1])
            except Exception as exc: messagebox.showerror(APP_TITLE, f"Не удалось открыть результат:\n{exc}")

    def open_log(self):
        LOG_FILE.touch(exist_ok=True)
        try: self._open_path(LOG_FILE)
        except Exception as exc: messagebox.showerror(APP_TITLE, f"Не удалось открыть журнал:\n{exc}")

    def open_output_folder(self):
        if not self.last_outputs:
            return
        folder = str(self.last_outputs[-1].parent)
        try:
            if sys.platform.startswith("win"):
                os.startfile(folder)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                import subprocess
                subprocess.Popen(["open", folder])
            else:
                import subprocess
                subprocess.Popen(["xdg-open", folder])
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Не удалось открыть папку:\n{e}")


if __name__ == "__main__":
    try:
        if "--self-test" in sys.argv:
            database = resource_path("mtr_data.json.gz")
            checker = Anonymizer(database)
            if not database.is_file() or not checker.registry or not checker.version:
                raise RuntimeError("Runtime-база не загружена")
            message = f"SELF_TEST_OK version={APP_VERSION} database={database.name} registry={len(checker.registry)}"
            # A windowed PyInstaller executable has no stdout on Windows.
            # The marker gives CI durable evidence that the packaged EXE itself
            # started and loaded its embedded production database.
            if getattr(sys, "frozen", False):
                # CI starts the windowed EXE with its extracted release folder
                # as the working directory.  Using cwd avoids PyInstaller
                # one-file path ambiguities and puts the marker beside the EXE.
                (Path.cwd() / "SELF_TEST_OK.txt").write_text(message, encoding="utf-8")
            else:
                print(message)
            raise SystemExit(0)
        if CONFIG_FILE.exists():
            configured = json.loads(CONFIG_FILE.read_text(encoding="utf-8")).get("tessdata")
            if configured: os.environ["TESSDATA_PREFIX"] = configured
        app = App()
        app.mainloop()
    except Exception as exc:
        APP_DIR.mkdir(parents=True, exist_ok=True)
        with LOG_FILE.open("a", encoding="utf-8") as handle:
            handle.write(f"{datetime.now():%Y-%m-%d %H:%M:%S} КРИТИЧЕСКАЯ ОШИБКА\n{traceback.format_exc()}\n")
        root = tk.Tk(); root.withdraw()
        messagebox.showerror(APP_TITLE, f"Критическая ошибка: {exc}\n\nПодробности: {LOG_FILE}")
